#!/usr/bin/env python
"""
Extract a CIS Benchmark PDF into a CISO Assistant v2 framework Excel.

Parses only the "Appendix: Summary Table" (recommendation numbers, titles and
their Automated/Manual tag) — no audit/remediation/rationale content, so the
output stays within the same IP boundary as the existing
cis-benchmark-kubernetes library.

Output: 5-sheet xlsx (library_meta, reco_meta, reco_content, cat_meta,
cat_content) in the exact shape of tools/excel/cis/cis-benchmark-kubernetes.xlsx,
ready for tools/convert_library_v2.py.

Usage:
    python extract_cis_benchmark.py <benchmark.pdf> [--ref-id cis-benchmark-xxx]
        [--name "..."] [--output out.xlsx] [--keep-empty-sections]
"""

import argparse
import datetime
import re
import sys
from pathlib import Path

import fitz  # pymupdf
from openpyxl import Workbook

SUMMARY_HEADING = "Appendix: Summary Table"
NAME_MAX_LENGTH = 200  # DB limit enforced by convert_library_v2.py
TABLE_HEADER_LINES = {"CIS Benchmark Recommendation", "Set", "Correctly", "Yes", "No"}
REF_ID_RE = re.compile(r"^\d+(\.\d+)*$")
TAG_RE = re.compile(r"\s*\((Automated|Manual)\)$")
VERSION_LINE_RE = re.compile(r"^[vV](\d+(?:\.\d+)*)\s*[-–]\s*(\d{2}-\d{2}-\d{4})$")
PAGE_LINE_RE = re.compile(r"^Page \d+$")
CLASSIFICATION_RE = re.compile(r"^(Internal Only|TLP:).*", re.IGNORECASE)


def cover_metadata(doc):
    """Return (title, version, publication_date) from the cover page."""
    title_parts = []
    for page in doc[:3]:
        lines = [line.strip() for line in page.get_text().splitlines() if line.strip()]
        for line in lines:
            m = VERSION_LINE_RE.match(line)
            if m:
                version = m.group(1)
                date = datetime.datetime.strptime(m.group(2), "%m-%d-%Y").date()
                return " ".join(title_parts), version, date
            if not CLASSIFICATION_RE.match(line) and not PAGE_LINE_RE.match(line):
                title_parts.append(line)
    raise SystemExit(
        "Could not find 'vX.Y.Z - MM-DD-YYYY' version line on the cover page"
    )


def default_ref_id(title):
    slug = title.lower()
    slug = re.sub(r"^cis\s+", "", slug)
    slug = re.sub(r"\bbenchmark\b", "", slug)
    slug = re.sub(r"[^a-z0-9.]+", "-", slug).strip("-")
    return f"cis-benchmark-{slug}"


def summary_table_lines(doc):
    """Yield content lines of the summary table, page headers stripped."""
    start = None
    for i, page in enumerate(doc):
        lines = [line.strip() for line in page.get_text().splitlines()]
        if SUMMARY_HEADING in lines and "CIS Benchmark Recommendation" in lines:
            start = i
            break
    if start is None:
        raise SystemExit(f"'{SUMMARY_HEADING}' page not found")

    for page in doc[start:]:
        for raw in page.get_text().splitlines():
            line = raw.strip()
            if not line:
                continue
            if line.startswith("Appendix:") and line != SUMMARY_HEADING:
                return
            if (
                line == SUMMARY_HEADING
                or line in TABLE_HEADER_LINES
                or PAGE_LINE_RE.match(line)
                or CLASSIFICATION_RE.match(line)
            ):
                continue
            # checkbox glyphs (o, private-use chars, wingdings squares)
            if len(line) <= 3 and not any(c.isalnum() for c in line):
                continue
            if line == "o":
                continue
            yield line


def is_successor(prev, cur):
    """Whether ref_id `cur` can follow `prev` in a depth-first numbering."""
    if prev is None:
        return len(cur) == 1
    if cur == prev + [1]:  # first child
        return True
    for k in range(1, len(prev) + 1):  # next sibling of prev or of an ancestor
        if cur == prev[: k - 1] + [prev[k - 1] + 1]:
            return True
    return False


FRAGMENT_RE = re.compile(r"^[\d.]+$")


def parse_rows(lines):
    """Return (rows, warnings). rows = [(ref_id, depth, name, tag_or_None)].

    Long ref_ids can wrap across lines inside the PDF's narrow ref column
    (e.g. '18.10.10.1.1' + '0' for 18.10.10.1.10), so consecutive numeric
    fragments are buffered and reassembled; a candidate is only accepted as a
    ref_id when it forms a valid depth-first successor of the previous one.
    """
    rows, warnings = [], []
    prev_parts = None
    cur_ref, cur_title = None, []
    pending = []  # consecutive numeric fragments not yet resolved

    def flush_row():
        nonlocal cur_ref, cur_title
        if cur_ref is None:
            return
        title = ""
        for part in cur_title:
            if title.endswith("-"):
                title += part
            elif title:
                title += " " + part
            else:
                title = part
        # short rows keep their checkbox glyph(s) on the title line: "... (Manual) o"
        title = re.sub(r"(\s+[o\u2610-\u2612\uE000-\uF8FF])+\s*$", "", title)
        m = TAG_RE.search(title)
        tag = m.group(1) if m else None
        if m:
            title = title[: m.start()].strip()
        if not title:
            warnings.append(f"empty title for ref_id {cur_ref}")
        rows.append((cur_ref, cur_ref.count(".") + 1, title, tag))

    def pending_to_title():
        for frag in pending:
            if frag.count(".") >= 2:
                warnings.append(
                    f"ref-like line '{frag}' treated as title text (after {cur_ref})"
                )
        cur_title.extend(pending)
        pending.clear()

    for line in lines:
        if FRAGMENT_RE.match(line):
            pending.append(line)
            combined = "".join(pending)
            candidates = [combined] if len(pending) > 1 else []
            candidates.append(line)
            matched = None
            for cand in candidates:
                if REF_ID_RE.match(cand) and is_successor(
                    prev_parts, [int(p) for p in cand.split(".")]
                ):
                    matched = cand
                    break
            if matched:
                if matched == line and len(pending) > 1:
                    pending.pop()
                    pending_to_title()  # earlier fragments belong to the previous title
                else:
                    pending.clear()
                flush_row()
                cur_ref, cur_title = matched, []
                prev_parts = [int(p) for p in matched.split(".")]
            continue
        pending_to_title()
        cur_title.append(line)
    pending_to_title()
    flush_row()

    seen = set()
    for ref, _, _, _ in rows:
        if ref in seen:
            warnings.append(f"duplicate ref_id {ref}")
        seen.add(ref)
    return rows, warnings


def prune_empty_sections(rows):
    """Drop non-assessable nodes with no assessable descendants (e.g. 'Introduction')."""
    kept = set()
    for ref, _, _, tag in rows:
        if tag:
            parts = ref.split(".")
            for k in range(1, len(parts) + 1):
                kept.add(".".join(parts[:k]))
    return [r for r in rows if r[0] in kept]


def write_xlsx(path, ref_id, name, description, pub_date, rows):
    wb = Workbook()

    ws = wb.active
    ws.title = "library_meta"
    for row in [
        ("type", "library"),
        ("urn", f"urn:intuitem:risk:library:{ref_id}"),
        ("version", "1"),
        ("publication_date", datetime.datetime.combine(pub_date, datetime.time())),
        ("locale", "en"),
        ("ref_id", ref_id),
        ("name", name),
        ("description", description),
        ("copyright", "© CIS Security"),
        ("provider", "CIS"),
        ("packager", "intuitem"),
    ]:
        ws.append(row)

    ws = wb.create_sheet("reco_meta")
    for row in [
        ("type", "framework"),
        ("base_urn", f"urn:intuitem:risk:req_node:{ref_id}"),
        ("urn", f"urn:intuitem:risk:framework:{ref_id}"),
        ("ref_id", ref_id),
        ("name", name),
        ("description", description),
        ("implementation_groups_definition", "cat"),
    ]:
        ws.append(row)

    ws = wb.create_sheet("reco_content")
    ws.append(
        (
            "assessable",
            "ref_id",
            "depth",
            "name",
            "description",
            "implementation_groups",
        )
    )
    for ref, depth, title, tag in rows:
        description = None
        if len(title) > NAME_MAX_LENGTH:
            description = title
            title = title[: NAME_MAX_LENGTH - 1].rsplit(" ", 1)[0] + "…"
        ws.append(
            (
                "x" if tag else None,
                ref,
                depth,
                title,
                description,
                tag[0] if tag else None,
            )
        )

    ws = wb.create_sheet("cat_meta")
    ws.append(("type", "implementation_groups"))
    ws.append(("name", "cat"))

    ws = wb.create_sheet("cat_content")
    ws.append(("ref_id", "name", "description"))
    ws.append(("A", "Automatic", None))
    ws.append(("M", "Manual", None))

    wb.save(path)


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("pdf", type=Path)
    parser.add_argument(
        "--ref-id", help="library/framework ref_id (default: derived from title)"
    )
    parser.add_argument(
        "--name", help="library/framework name (default: cover page title)"
    )
    parser.add_argument(
        "--output", "-o", type=Path, help="output xlsx (default: <ref_id>.xlsx)"
    )
    parser.add_argument(
        "--keep-empty-sections",
        action="store_true",
        help="keep sections that contain no assessable recommendation (e.g. 'Introduction')",
    )
    parser.add_argument(
        "--rename",
        action="append",
        default=[],
        metavar="REF=TITLE",
        help="errata override: replace the extracted title of REF (repeatable); "
        "use when the PDF's summary table contradicts the recommendation body",
    )
    args = parser.parse_args()

    doc = fitz.open(args.pdf)
    title, version, pub_date = cover_metadata(doc)
    name = args.name or title
    ref_id = args.ref_id or default_ref_id(title)
    description = f"{name} v{version}"

    rows, warnings = parse_rows(summary_table_lines(doc))
    doc.close()
    bad = [r for r in args.rename if "=" not in r]
    if bad:
        sys.exit(f"--rename expects REF=New title, got: {', '.join(bad)}")
    renames = dict(r.split("=", 1) for r in args.rename)
    if renames:
        rows = [(ref, d, renames.pop(ref, t), tag) for ref, d, t, tag in rows]
        for ref in renames:
            warnings.append(f"--rename ref {ref} not found in extracted rows")
    if not args.keep_empty_sections:
        before = len(rows)
        rows = prune_empty_sections(rows)
        if before != len(rows):
            print(
                f"pruned {before - len(rows)} section(s) with no assessable recommendation"
            )

    prev_depth = 0
    for ref, depth, _, _ in rows:
        if depth > prev_depth + 1:
            warnings.append(f"depth jump at {ref} ({prev_depth} -> {depth})")
        prev_depth = depth

    output = args.output or Path(f"{ref_id}.xlsx")
    write_xlsx(output, ref_id, name, description, pub_date, rows)

    n_auto = sum(1 for r in rows if r[3] == "Automated")
    n_manual = sum(1 for r in rows if r[3] == "Manual")
    n_section = sum(1 for r in rows if r[3] is None)
    print(f"{output}: {name} v{version} ({pub_date})")
    print(
        f"  {len(rows)} nodes: {n_section} sections, {n_auto} automated, {n_manual} manual"
    )
    for w in warnings:
        print(f"  WARNING: {w}", file=sys.stderr)
    if warnings:
        sys.exit(2)


if __name__ == "__main__":
    main()
