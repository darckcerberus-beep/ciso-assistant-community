"""
Simple script to convert a TISAX Excel workbook into a CISO Assistant Excel file.
"""

import openpyxl
import re
import argparse

# Keep the version in one place so a future TISAX release only requires this
# constant to be updated for the generated identifiers and filename.
TISAX_VERSION = "v2027"
TISAX_LABEL = f"TISAX {TISAX_VERSION}"

parser = argparse.ArgumentParser(
    prog="convert_tisax",
    description=f"convert TISAX controls official {TISAX_VERSION} Excel file to CISO Assistant Excel file",
)

parser.add_argument("filename", help="name of official TISAX Excel file")
args = parser.parse_args()
input_file_name = args.filename
output_file_name = f"tisax-{TISAX_VERSION}.xlsx"

LIBRARY_COPYRIGHT = """© 2026 ENX Association, an Association according to the French Law of 1901, registered under No. w923004198 at the Sous-préfecture of Boulogne-Billancourt, France.
This work of ENX's Working Group ISA was provided to the VDA in the present version by the ENX Association for published by the VDA as the VDA ISA. It is made to all interested parties free of charge under the following licensing terms. The release in the VDA is done by the VDA's Working Group Information Security and Economic Protection. Publication takes place with the consent of the rights holder. The VDA is responsible for the publication of the VDA ISA.
The Tab ""Data Protection"" is provided, owned and copyrighted by VERBAND DER AUTOMOBILINDUSTRIE e.V. (VDA, German Association of the Automotive Industry); Behrenstr. 35; 10117 Berlin"
This work has been licensed under Creative Commons Attribution - No Derivative Works 4.0 International Public License. In addition, You are granted the right to distribute derivatives under certain terms as detailed in section 9 which are not part of the Creative Commons license. The complete and valid text of the license is to be found in line 17ff."""
PACKAGER = "intuitem"

LIBRARY_DESCRIPTION = """ISA provides the basis for
- a self-assessment to determine the state of information security in an organization (e.g. company)
- audits performed by internal departments (e.g. Internal Audit, Information Security)
- TISAXⓇ Assessments (Trusted Information Security Assessment Exchange, https://enx.com/tisax/)
Source: https://enx.com/en-US/TISAX/downloads/"""

LIBRARY_URN = f"urn:{PACKAGER}:risk:library:tisax-{TISAX_VERSION}"
FRAMEWORK_BASE_URN = f"urn:{PACKAGER}:risk:req_node:tisax-{TISAX_VERSION}"
FRAMEWORK_URN = f"urn:{PACKAGER}:risk:framework:tisax-{TISAX_VERSION}"
TISAX_NAME = f"Trusted Information Security Assessment Exchange (TISAX) {TISAX_VERSION}"
SCORES_SHEET = "scores"
IMPLEMENTATION_GROUPS_SHEET = "implementation_groups"

LIBRARY_META_ROWS = [
    ["type", "library"],
    ["urn", LIBRARY_URN],
    ["version", 1],
    ["locale", "en"],
    ["ref_id", TISAX_LABEL],
    ["name", TISAX_NAME],
    ["description", LIBRARY_DESCRIPTION],
    ["copyright", LIBRARY_COPYRIGHT],
    ["provider", "VDA"],
    ["packager", PACKAGER],
]

CONTROLS_META_ROWS = [
    ["type", "framework"],
    ["base_urn", FRAMEWORK_BASE_URN],
    ["urn", FRAMEWORK_URN],
    ["ref_id", TISAX_LABEL],
    ["name", TISAX_NAME],
    ["description", LIBRARY_DESCRIPTION],
    ["min_score", "0"],
    ["max_score", "5"],
    ["scores_definition", SCORES_SHEET],
    ["implementation_groups_definition", IMPLEMENTATION_GROUPS_SHEET],
]

SCORES_META_ROWS = [
    ["type", "scores"],
    ["name", SCORES_SHEET],
]

SCORES_CONTENT_ROWS = [
    ["score", "name", "description"],
    [
        0,
        "Incomplete",
        "A process is not available, not followed or not suitable for achieving the objective.",
    ],
    [
        1,
        "Performed",
        "An undocumented or incompletely documented process is followed and indicators exist that it achieves its objective.",
    ],
    [
        2,
        "Managed",
        "A process achieving its objectives is followed. Process documentation and process implementation evidence are available.",
    ],
    [
        3,
        "Established",
        "A standard process integrated into the overall system is followed. Dependencies on other processes are documented and suitable interfaces are created. Evidence exists that the process has been used sustainably and actively over an extended period.",
    ],
    [
        4,
        "Predictable",
        "An established process is followed. The effectiveness of the process is continually monitored by collecting key figures. Limit values are defined at which the process is considered to be insufficiently effective and requires adjustment. (Key Performance Indicators)",
    ],
    [
        5,
        "Optimizing",
        "A predictable process with continual improvement as a major objective is followed. Improvement is actively advanced by dedicated resources.",
    ],
]

IMPLEMENTATION_GROUPS_META_ROWS = [
    ["type", "implementation_groups"],
    ["name", IMPLEMENTATION_GROUPS_SHEET],
]

IMPLEMENTATION_GROUPS_CONTENT_ROWS = [
    ["ref_id", "name", "description"],
    ["must", "Requirements (must)", "Strict requirements without any exemptions."],
    [
        "should",
        "Requirements (should)",
        "Must be implemented by the organization. In certain circumstances, however, there may be a valid justification for non-compliance with these requirements. In case of any deviation, its effects must be understood by the organization and it must be plausibly justified.",
    ],
    [
        "high",
        "In case of high protection needs",
        "Must additionally be met if the tested subject has high protection needs.",
    ],
    [
        "very_high",
        "In case of very high protection needs",
        "Must additionally be met if the tested subject has very high protection needs.",
    ],
    [
        "SGA",
        "For Simplified Group Assessments (SGA)",
        "A simplified way to audit very large organizations with a high maturity. An example is the TISAX Simplified Group Assessment mechanism that is an option for TISAX Assessments of an assessment scope with a large number of sites.",
    ],
    [
        "vehicle",
        "For vehicles classified as requiring protection",
        "Protects physical prototypes which are classified as requiring protection. Prototypes include vehicles, components and parts. The owner of the intellectual property for the prototype is considered the owner of the prototype. The owner's commissioning department is responsible for classifying the protection need of a prototype. For prototypes classified as requiring high or very high protection, the minimum requirements for prototype protection must be applied.",
    ],
]


def is_effective_value(value):
    """Return False for empty cells and textual None placeholders."""
    if value is None:
        return False

    text = str(value).replace("\u200b", "").strip()
    return bool(text) and text.casefold() != "none"


def effective_text(value):
    """Return usable cell text, excluding None placeholders."""
    if not is_effective_value(value):
        return ""
    return str(value).replace("\u200b", "").strip()


def split_requirement_items(requirement):
    """Split a requirement into items, starting a new item at each '+' line."""
    if not is_effective_value(requirement):
        return []

    lines = [line.replace("\u200b", "") for line in str(requirement).splitlines()]
    if not any(line.lstrip().startswith("+") for line in lines):
        text = "\n".join(lines).strip()
        return [text] if is_effective_value(text) else []

    items = []
    current_item = []
    for line in lines:
        if line.lstrip().startswith("+"):
            if current_item:
                items.append("\n".join(current_item).strip())
            current_item = [line.rstrip()]
        elif current_item:
            current_item.append(line.rstrip())
        elif line.strip():
            current_item = [line.rstrip()]

    if current_item:
        items.append("\n".join(current_item).strip())

    return [item for item in items if item]


def append_requirement_nodes(rows, depth, label, requirement, implementation_group):
    """Add a non-assessable requirement label and its assessable child nodes."""
    items = split_requirement_items(requirement)
    if not items:
        return

    rows.append(("", depth, "", label, "", ""))
    for item in items:
        rows.append(("x", depth + 1, "", "", item, implementation_group))

print(f"▶️  Parsing \"{input_file_name}\"")

# Define variable to load the dataframe
dataframe = openpyxl.load_workbook(input_file_name)
output_table = []

for tab in dataframe:
    print(f"⌛ Parsing tab \"{tab.title}\"...")
    title = tab.title
    if title in ("Information Security", "Prototype Protection", "Data Protection"):
        for row in tab:
            req_should = None
            req_very_high = None
            req_high = None
            req_sga = None
            req_vehicle = None
            further_info = None
            ex_normal = None
            ex_high = None
            ex_very_high = None
            if title == "Information Security":
                (
                    _,
                    _,
                    control_number,
                    _,
                    _,
                    _,
                    _,
                    control_question,
                    objective,
                    req_must,
                    req_should,
                    req_high,
                    req_very_high,
                    req_sga,
                    _,
                    _,
                    _,
                    _,
                    _,
                    _,
                    _,
                    _,
                    further_info,
                    ex_normal,
                    ex_high,
                    ex_very_high,
                ) = (r.value for r in row[0:26])
            elif title == "Prototype Protection":
                (
                    _,
                    _,
                    control_number,
                    _,
                    _,
                    _,
                    _,
                    control_question,
                    objective,
                    req_must,
                    req_should,
                _,
                ) = (r.value for r in row[0:12])
                # In the v2027 workbook, "Further information" is in column R.
                further_info = row[17].value
            elif title == "Data Protection":
                (
                    _,
                    _,
                    control_number,
                    _,
                    _,
                    _,
                    _,
                    control_question,
                    objective,
                    req_must,
                ) = (r.value for r in row[0:10])
                # In the v2027 workbook, "Further information" is in column R.
                further_info = row[17].value
            if type(control_number) == int:
                control_number = str(control_number)
            if control_number and re.fullmatch(r"\d", control_number):
                level = 2
                print(control_number, control_question)
                output_table.append(
                    ("", 1, control_number, control_question, effective_text(objective), "")
                )
            if control_number and re.fullmatch(r"\d\.\d+", control_number):
                level = 3
                print(control_number, control_question)
                output_table.append(
                    ("", 2, control_number, control_question, effective_text(objective), "")
                )
            if control_number and re.fullmatch(r"\d\.\d+\.\d+", control_number):
                if re.match(r"Superseded by", control_question):
                    print(f"⏭️  Skipping \"{control_number}\"")
                # print(control_number, control_question)
                output_table.append(("", level, control_number, control_question, effective_text(objective), ""))
                append_requirement_nodes(
                    output_table, level + 1, "(must)", req_must, "must"
                )
                if is_effective_value(req_should):
                    append_requirement_nodes(
                        output_table, level + 1, "(should)", req_should, "should"
                    )
                if is_effective_value(req_high):
                    append_requirement_nodes(
                        output_table,
                        level + 1,
                        "(for high protection needs)",
                        req_high,
                        "high",
                    )
                if is_effective_value(req_very_high):
                    append_requirement_nodes(
                        output_table,
                        level + 1,
                        "(for very high protection needs)",
                        req_very_high,
                        "very_high",
                    )
                if is_effective_value(req_sga):
                    append_requirement_nodes(
                        output_table,
                        level + 1,
                        "(for Simplified Group Assessments)",
                        req_sga,
                        "SGA",
                    )
                if is_effective_value(req_vehicle):
                    append_requirement_nodes(
                        output_table,
                        level + 1,
                        "(for vehicles classified as requiring protection)",
                        req_vehicle,
                        "vehicle",
                    )
                if is_effective_value(further_info):
                    output_table.append(
                        ("", level + 1, "", "Further information", further_info, "")
                    )

print(f"⌛ Generating \"{output_file_name}\"")
wb_output = openpyxl.Workbook()
wb_output.active.title = "library_meta"

for row in LIBRARY_META_ROWS:
    wb_output["library_meta"].append(row)

controls_meta = wb_output.create_sheet("controls_meta")
for row in CONTROLS_META_ROWS:
    controls_meta.append(row)

controls_content = wb_output.create_sheet("controls_content")
controls_content.append(
    ["assessable", "depth", "ref_id", "name", "description", "implementation_groups"]
)
for row in output_table:
    controls_content.append(row)

scores_meta = wb_output.create_sheet("scores_meta")
for row in SCORES_META_ROWS:
    scores_meta.append(row)

scores_content = wb_output.create_sheet("scores_content")
for row in SCORES_CONTENT_ROWS:
    scores_content.append(row)

implementation_groups_meta = wb_output.create_sheet("implementation_groups_meta")
for row in IMPLEMENTATION_GROUPS_META_ROWS:
    implementation_groups_meta.append(row)

implementation_groups_content = wb_output.create_sheet("implementation_groups_content")
for row in IMPLEMENTATION_GROUPS_CONTENT_ROWS:
    implementation_groups_content.append(row)

print(f"✅ Generated \"{output_file_name}\"")
wb_output.save(output_file_name)
