"""Audit the contents of the Excel tools directory.

This script is mainly intended to identify Excel workbooks that are not useful
for the project and can potentially be removed. It also reports Python files,
framework build instructions, OS-specific custom scripts, libraries, and all
other file types so the directory can be reviewed before cleanup.

The report is written both to the console and to a log file. Excel files are
checked for the required `library_meta` sheet, while every other file type is
grouped into a dedicated report section.
"""

from __future__ import annotations

import argparse
from collections import defaultdict
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from tqdm import tqdm


REQUIRED_SHEET = "library_meta"
EXCEL_EXTENSIONS = {".xlsx", ".xlsm"}
PYTHON_EXTENSIONS = {".py"}
PYTHON_FILE_NAMES = {"requirements.txt"}
FRAMEWORK_BUILD_INSTRUCTION_NAMES = {"readme.md"}
OS_SPECIFIC_SCRIPT_EXTENSIONS = {".sh", ".bat", ".ps1"}
LIBRARY_EXTENSIONS = {".yaml"}
DEFAULT_EXCEL_DIR = Path(__file__).resolve().parent / "excel"
DEFAULT_LOG_FILE = Path(__file__).resolve().parent / "check_excel_folder_contents.log"


def find_files(directory: Path) -> Iterable[Path]:
    """Yield files below *directory*, excluding temporary Excel lock files."""
    for path in sorted(directory.rglob("*")):
        if path.is_file() and not path.name.startswith("~$"):
            yield path


def get_file_type(path: Path) -> str:
    """Return a display label for a file extension."""
    return path.suffix.lower() or "[no extension]"


def has_required_sheet(path: Path) -> bool:
    """Return whether an Excel file contains a sheet named `library_meta`."""
    keep_vba = path.suffix.lower() == ".xlsm"
    workbook = load_workbook(
        path,
        read_only=True,
        data_only=True,
        keep_links=False,
        keep_vba=keep_vba,
    )
    try:
        return REQUIRED_SHEET in workbook.sheetnames
    finally:
        workbook.close()


def append_section(lines: list[str], title: str, paths: Iterable[object]) -> None:
    """Append one consistently formatted report section."""
    paths = list(paths)
    lines.append(f"=== {title} ({len(paths)}) ===")
    if paths:
        lines.extend(str(path) for path in paths)
    else:
        lines.append("ℹ️  None found.")
    lines.append("")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Audit Excel, Python, and other files in the tools/excel directory. "
            "Excel files are checked for a library_meta sheet."
        )
    )
    parser.add_argument(
        "--excel-dir",
        type=Path,
        default=DEFAULT_EXCEL_DIR,
        help=f"Directory to scan (default: {DEFAULT_EXCEL_DIR}).",
    )
    parser.add_argument(
        "--log-file",
        type=Path,
        default=DEFAULT_LOG_FILE,
        help=f"Log file to write (default: {DEFAULT_LOG_FILE}).",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    excel_dir = args.excel_dir.resolve()
    log_file = args.log_file.resolve()

    missing_library_meta: list[Path] = []
    python_files: list[Path] = []
    framework_build_instructions: list[Path] = []
    os_specific_scripts: list[Path] = []
    library_files: list[Path] = []
    other_files: defaultdict[str, list[Path]] = defaultdict(list)
    errors: list[str] = []
    excel_file_count = 0
    valid_excel_file_count = 0

    if not excel_dir.is_dir():
        errors.append(f"❌ [ERROR] Excel directory not found: {excel_dir}")
    else:
        files_to_scan = list(find_files(excel_dir))
        for path in tqdm(
            files_to_scan,
            total=len(files_to_scan),
            desc="Scanning files",
            unit="file",
        ):
            file_type = get_file_type(path)

            if file_type in EXCEL_EXTENSIONS:
                excel_file_count += 1
                try:
                    if has_required_sheet(path):
                        valid_excel_file_count += 1
                    else:
                        missing_library_meta.append(path)
                except Exception as error:  # openpyxl can raise file-specific errors.
                    errors.append(f"❌ [ERROR] {path}: {error}")
            elif file_type in PYTHON_EXTENSIONS or path.name.lower() in PYTHON_FILE_NAMES:
                python_files.append(path)
            elif path.name.lower() in FRAMEWORK_BUILD_INSTRUCTION_NAMES:
                framework_build_instructions.append(path)
            elif file_type in OS_SPECIFIC_SCRIPT_EXTENSIONS:
                os_specific_scripts.append(path)
            elif file_type in LIBRARY_EXTENSIONS:
                library_files.append(path)
            else:
                other_files[file_type].append(path)

    other_file_count = sum(len(paths) for paths in other_files.values())
    lines = [
        f"📂 Scanned directory: {excel_dir}",
        "",
        "📊 === Summary ===",
        f"📗 Excel files checked: {excel_file_count}",
        f"✅ Excel files with \"{REQUIRED_SHEET}\": {valid_excel_file_count}",
        f"⚠️  Excel files without \"{REQUIRED_SHEET}\": {len(missing_library_meta)}",
        f"🐍 Python files found: {len(python_files)}",
        f"📘 Framework Build Instructions found: {len(framework_build_instructions)}",
        f"🖥️  OS-specific custom scripts found: {len(os_specific_scripts)}",
        f"📚 Libraries found: {len(library_files)}",
        f"📁 Other files found: {other_file_count}",
    ]

    if other_files:
        lines.extend(
            f"   - {file_type}: {len(paths)}"
            for file_type, paths in sorted(other_files.items())
        )
    else:
        lines.append("ℹ️  No other file types found.")
    lines.append("")

    append_section(
        lines,
        f'⚠️  Excel files without a "{REQUIRED_SHEET}" sheet',
        missing_library_meta,
    )
    append_section(lines, "🐍 Python files (.py, requirements.txt)", python_files)
    append_section(
        lines,
        "📘 Framework Build Instructions (README.md)",
        framework_build_instructions,
    )
    append_section(
        lines,
        "🖥️  OS-specific custom scripts (.sh, .bat, .ps1)",
        os_specific_scripts,
    )
    append_section(lines, "📚 Libraries (.yaml)", library_files)

    for file_type, paths in sorted(other_files.items()):
        append_section(lines, f"📁 Other files of type \"{file_type}\"", paths)

    append_section(lines, "❌ Errors", errors)

    log_file.parent.mkdir(parents=True, exist_ok=True)
    log_file.write_text("\n".join(lines).rstrip() + "\n", encoding="utf-8")

    print("\n".join(lines).rstrip())
    return 1 if missing_library_meta or other_files or errors else 0


if __name__ == "__main__":
    raise SystemExit(main())
