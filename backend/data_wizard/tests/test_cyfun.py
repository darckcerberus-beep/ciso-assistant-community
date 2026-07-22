import io
from pathlib import Path

import openpyxl
import pytest

from core.models import (
    ComplianceAssessment,
    LoadedLibrary,
    RequirementAssessment,
    StoredLibrary,
)
from data_wizard.cyfun_helpers import (
    CYFUN_FRAMEWORK_URN,
    CYFUN_LIBRARY_URN,
    FUNCTION_SHEETS,
    process_cyfun_file,
)

URL = "/api/data-wizard/load-file/"

HEADERS = [
    "Category",
    "Controls linked to the management aspects",
    "Key Measure",
    "Subcategory",
    "Assurance level",
    "Requirement",
    "Documentation Score",
    "Implementation Score",
    "Subcategory Documentation Maturity Score",
    "Subcategory Implementation Maturity Score",
    "Category Documentation Maturity Score",
    "Category Implementation Maturity Score",
    "Comments and/or additional information",
    "Assessor comments",
]


# The BASIC edition has no "Assurance level" column: requirement/scores/comments
# all sit one column left, and the level marker is above "Documentation Score".
HEADERS_BASIC = [h for h in HEADERS if h != "Assurance level"]


def build_basic_workbook(rows_by_sheet: dict, level: str = "BASIC") -> bytes:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for name in FUNCTION_SHEETS:
        ws = wb.create_sheet(name)
        ws.cell(1, 6, level)
        for col, header in enumerate(HEADERS_BASIC, 1):
            ws.cell(2, col, header)
        for row_idx, row in enumerate(rows_by_sheet.get(name, []), 3):
            for col, value in row.items():
                ws.cell(row_idx, col, value)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_workbook(rows_by_sheet: dict, level: str = "ESSENTIAL") -> bytes:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for name in FUNCTION_SHEETS:
        ws = wb.create_sheet(name)
        ws.cell(1, 7, level)
        for col, header in enumerate(HEADERS, 1):
            ws.cell(2, col, header)
        for row_idx, row in enumerate(rows_by_sheet.get(name, []), 3):
            for col, value in row.items():
                ws.cell(row_idx, col, value)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class TestProcessCyfunFile:
    def test_parses_scores_and_comments(self):
        content = build_workbook(
            {
                "GOVERN": [
                    {
                        5: "Important",
                        6: "GV.OC-01.1: The organisation's mission shall be established.",
                        7: 3,
                        8: 2,
                        13: "our comment",
                        14: "assessor note",
                    }
                ],
                "PROTECT": [
                    {
                        5: "Essential",
                        6: "PR.AA-01.1: Identities are managed.",
                        7: 4,
                        8: 4,
                    }
                ],
            }
        )
        parsed = process_cyfun_file(content)
        assert parsed["assurance_level"] == "essential"
        assert len(parsed["records"]) == 2
        record = parsed["records"][0]
        assert record["ref_id"] == "GV.OC-01.1"
        assert record["documentation_score"] == 3
        assert record["implementation_score"] == 2
        assert (
            record["observations"] == "our comment\n\nAssessor comments: assessor note"
        )

    def test_na_maps_to_not_applicable(self):
        content = build_workbook(
            {"GOVERN": [{6: "GV.OC-02.1: Stakeholders.", 7: "N/A", 8: "N/A"}]}
        )
        record = process_cyfun_file(content)["records"][0]
        assert record["compliance_result"] == "not_applicable"
        assert "implementation_score" not in record
        assert "documentation_score" not in record

    def test_skips_non_requirement_rows(self):
        content = build_workbook(
            {
                "GOVERN": [
                    {6: "NO REQUIREMENT / Guidance to be considered"},
                    {6: None},
                    {6: "GV.RM-01.1: Objectives.", 7: 1, 8: 1},
                ]
            }
        )
        parsed = process_cyfun_file(content)
        assert [r["ref_id"] for r in parsed["records"]] == ["GV.RM-01.1"]

    def test_level_falls_back_to_rows_when_g1_missing(self):
        content = build_workbook(
            {
                "GOVERN": [
                    {5: "Basic", 6: "GV.OC-03.1: Legal requirements.", 7: 2, 8: 2},
                    {5: "Important", 6: "GV.OC-01.1: Mission.", 7: 2, 8: 2},
                ]
            },
            level="",
        )
        assert process_cyfun_file(content)["assurance_level"] == "important"

    def test_parses_basic_edition_layout(self):
        content = build_basic_workbook(
            {
                "GOVERN": [
                    {
                        5: "GV.OC-03.1: Legal requirements shall be identified.",
                        6: 2,
                        7: 3,
                        12: "tracked in legal register",
                        13: "verify register completeness",
                    },
                    {5: "GV.RM-03.1: Risk strategy.", 6: "N/A", 7: "N/A"},
                ],
                "IDENTIFY": [
                    {
                        5: "ID.AM-5.1: Unpadded ref as in the real BASIC tool.",
                        6: 1,
                        7: 2,
                    }
                ],
            }
        )
        parsed = process_cyfun_file(content)
        assert parsed["assurance_level"] == "basic"
        assert len(parsed["records"]) == 3
        assert parsed["records"][2]["ref_id"] == "ID.AM-05.1"
        record = parsed["records"][0]
        assert record["ref_id"] == "GV.OC-03.1"
        assert record["documentation_score"] == 2
        assert record["implementation_score"] == 3
        assert (
            record["observations"]
            == "tracked in legal register\n\nAssessor comments: verify register completeness"
        )
        assert parsed["records"][1]["compliance_result"] == "not_applicable"

    def test_rejects_workbook_with_unrecognized_sheet_headers(self):
        content = build_workbook(
            {"GOVERN": [{6: "GV.RM-01.1: Objectives.", 7: 1, 8: 1}]}
        )
        wb = openpyxl.load_workbook(io.BytesIO(content))
        wb["PROTECT"].cell(2, 6, "Exigence")
        buf = io.BytesIO()
        wb.save(buf)
        with pytest.raises(ValueError, match="UnrecognizedCyfunWorkbook"):
            process_cyfun_file(buf.getvalue())

    def test_rejects_2023_workbook(self):
        wb = openpyxl.Workbook()
        wb.active.title = "BASIC Details"
        wb.create_sheet("Introduction")
        buf = io.BytesIO()
        wb.save(buf)
        with pytest.raises(ValueError, match="CyFun2023WorkbookNotSupported"):
            process_cyfun_file(buf.getvalue())

    def test_rejects_unknown_workbook(self):
        wb = openpyxl.Workbook()
        wb.active.title = "Sheet1"
        buf = io.BytesIO()
        wb.save(buf)
        with pytest.raises(ValueError, match="UnrecognizedCyfunWorkbook"):
            process_cyfun_file(buf.getvalue())


@pytest.fixture
def cyfun_stored_library(db):
    if not StoredLibrary.objects.filter(urn=CYFUN_LIBRARY_URN).exists():
        path = (
            Path(__file__).resolve().parents[2]
            / "library"
            / "libraries"
            / "cyfun2025.yaml"
        )
        stored, error = StoredLibrary.store_library_content(path.read_bytes())
        assert error is None
        assert stored is not None


@pytest.mark.django_db
class TestCyfunEndpoint:
    def _post(self, client, content: bytes, folder_id, name=None):
        headers = {
            "HTTP_X_MODEL_TYPE": "CyFunAssessment",
            "HTTP_X_FOLDER_ID": str(folder_id),
            "HTTP_CONTENT_DISPOSITION": "attachment; filename=cyfun.xlsx",
            "content_type": "application/octet-stream",
        }
        if name:
            headers["HTTP_X_NAME"] = name
        return client.post(URL, data=content, **headers)

    def test_import_creates_scored_assessment(
        self, knox_admin_client, domain_folder, cyfun_stored_library
    ):
        content = build_workbook(
            {
                "GOVERN": [
                    {
                        5: "Important",
                        6: "GV.OC-01.1: The organisation's mission shall be established.",
                        7: 3,
                        8: 2,
                        13: "our comment",
                    },
                    {
                        5: "Essential",
                        6: "GV.OC-02.1: Stakeholders.",
                        7: "N/A",
                        8: "N/A",
                    },
                ]
            }
        )
        resp = self._post(
            knox_admin_client, content, domain_folder.id, name="CyFun test"
        )
        assert resp.status_code == 200, resp.json()
        results = resp.json()["results"]
        assert results["successful"] == 2
        assert results["failed"] == 0

        assert LoadedLibrary.objects.filter(urn=CYFUN_LIBRARY_URN).exists()
        ca = ComplianceAssessment.objects.get(name="CyFun test")
        assert ca.framework.urn == CYFUN_FRAMEWORK_URN
        assert ca.selected_implementation_groups == ["E"]
        assert ca.scoring_enabled
        assert ca.show_documentation_score
        assert ca.min_score == 1 and ca.max_score == 5
        assert ca.score_calculation_method == "average_of_averages"

        scored = RequirementAssessment.objects.get(
            compliance_assessment=ca, requirement__ref_id="GV.OC-01.1"
        )
        assert scored.documentation_score == 3
        assert scored.score == 2
        assert scored.is_scored
        assert scored.observation == "our comment"

        na = RequirementAssessment.objects.get(
            compliance_assessment=ca, requirement__ref_id="GV.OC-02.1"
        )
        assert na.result == "not_applicable"

    def test_unknown_workbook_reports_error(self, knox_admin_client, domain_folder):
        wb = openpyxl.Workbook()
        buf = io.BytesIO()
        wb.save(buf)
        resp = self._post(knox_admin_client, buf.getvalue(), domain_folder.id)
        assert resp.status_code == 200
        results = resp.json()["results"]
        assert results["failed"] == 1
        assert results["errors"][0]["error"] == "UnrecognizedCyfunWorkbook"
