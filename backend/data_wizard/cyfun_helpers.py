import io
import re

import openpyxl

CYFUN_LIBRARY_URN = "urn:intuitem:risk:library:ccb-cyfun2025"
CYFUN_FRAMEWORK_URN = "urn:intuitem:risk:framework:ccb-cyfun2025"

FUNCTION_SHEETS = ("GOVERN", "IDENTIFY", "PROTECT", "DETECT", "RESPOND", "RECOVER")
LEVELS = ("basic", "important", "essential")
LEVEL_TO_GROUP = {"basic": "B", "important": "I", "essential": "E"}

REF_ID_PATTERN = re.compile(r"^[A-Z]{2}\.[A-Z]{2}-\d+(\.\d+)?$")

# The framework zero-pads subcategory numbers (ID.AM-05) but some workbook
# editions don't (ID.AM-5.1 in the BASIC tool).
SUBCATEGORY_PAD_PATTERN = re.compile(r"-(\d)(?=\.|$)")

HEADER_ROW = 2
DATA_START_ROW = 3


def _detect_columns(sheet):
    """Map header labels to column indices; the BASIC edition has no
    'Assurance level' column, shifting everything after 'Subcategory' left."""
    headers = {}
    for cell in sheet[HEADER_ROW]:
        if cell.value:
            headers[str(cell.value).strip().lower()] = cell.column
    required = ("requirement", "documentation score", "implementation score")
    if not all(key in headers for key in required):
        return None
    return {
        "level": headers.get("assurance level"),
        "requirement": headers["requirement"],
        "doc_score": headers["documentation score"],
        "impl_score": headers["implementation score"],
        "comments": headers.get("comments and/or additional information"),
        "assessor_comments": headers.get("assessor comments"),
    }


def _parse_score(value):
    """Returns (score, is_na)."""
    if value is None or value == "":
        return None, False
    if isinstance(value, str):
        if value.strip().upper() in ("N/A", "NA"):
            return None, True
        try:
            return int(float(value)), False
        except ValueError:
            return None, False
    if isinstance(value, (int, float)):
        return int(value), False
    return None, False


def _cell(row, col):
    if not col or len(row) < col:
        return None
    return row[col - 1]


def _cell_text(row, col):
    return str(_cell(row, col) or "").strip()


def process_cyfun_file(file_content: bytes) -> dict:
    workbook = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
    sheet_names = set(workbook.sheetnames)
    if any(name.endswith("Details") for name in sheet_names):
        raise ValueError("CyFun2023WorkbookNotSupported")
    if not all(name in sheet_names for name in FUNCTION_SHEETS):
        raise ValueError("UnrecognizedCyfunWorkbook")

    level = None
    levels_seen = set()
    records = []
    for sheet_name in FUNCTION_SHEETS:
        sheet = workbook[sheet_name]
        cols = _detect_columns(sheet)
        if cols is None:
            raise ValueError("UnrecognizedCyfunWorkbook")
        if level is None:
            marker = str(sheet.cell(1, cols["doc_score"]).value or "").strip().lower()
            if marker in LEVELS:
                level = marker
        for row in sheet.iter_rows(min_row=DATA_START_ROW, values_only=True):
            ref_id = (
                _cell_text(row, cols["requirement"]).split(":")[0].strip().rstrip(".")
            )
            if not REF_ID_PATTERN.match(ref_id):
                continue
            ref_id = SUBCATEGORY_PAD_PATTERN.sub(r"-0\1", ref_id)
            row_level = _cell_text(row, cols["level"]).lower()
            if row_level in LEVELS:
                levels_seen.add(row_level)

            record = {"ref_id": ref_id, "assessable": True}
            doc_score, doc_na = _parse_score(_cell(row, cols["doc_score"]))
            impl_score, impl_na = _parse_score(_cell(row, cols["impl_score"]))
            if doc_na or impl_na:
                record["compliance_result"] = "not_applicable"
            else:
                if doc_score is not None:
                    record["documentation_score"] = doc_score
                if impl_score is not None:
                    record["implementation_score"] = impl_score

            observations = _cell_text(row, cols["comments"])
            assessor_comments = _cell_text(row, cols["assessor_comments"])
            if assessor_comments:
                assessor_comments = f"Assessor comments: {assessor_comments}"
                observations = (
                    f"{observations}\n\n{assessor_comments}"
                    if observations
                    else assessor_comments
                )
            if observations:
                record["observations"] = observations
            records.append(record)

    if not records:
        raise ValueError("EmptyCyfunWorkbook")
    if level not in LEVELS:
        level = next((lvl for lvl in reversed(LEVELS) if lvl in levels_seen), None)
    return {"assurance_level": level, "records": records}
