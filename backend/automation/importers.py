import csv
import io
import json

RESULT_VALUES = {"pass", "fail", "not_applicable", "error", "not_checked"}

# OCSF Compliance Finding (class_uid 2003) status mapping.
# Warning = needs human review, Unknown = not measured -> both land on not_checked.
OCSF_STATUS_MAP = {
    "pass": "pass",
    "fail": "fail",
    "warning": "not_checked",
    "unknown": "not_checked",
}
OCSF_STATUS_ID_MAP = {1: "pass", 2: "not_checked", 3: "fail", 99: "not_checked"}


class ImportError_(Exception):
    @property
    def message(self):
        return self.args[0] if self.args else "import failed"


PARSE_ERROR_CAP = 20


def _record_parse_error(extras, item):
    if len(extras["parse_errors"]) < PARSE_ERROR_CAP:
        extras["parse_errors"].append(item)


def _entry(row, extras):
    ref_id = (row.get("ref_id") or "").strip()
    result = (row.get("result") or "").strip().lower()
    if not ref_id or result not in RESULT_VALUES:
        _record_parse_error(extras, row)
        return None
    return {
        "ref_id": ref_id,
        "result": result,
        "actual": (row.get("actual") or "").strip(),
        "expected": (row.get("expected") or "").strip(),
        "message": (row.get("message") or "").strip(),
    }


def parse_csv(file):
    extras = {"parse_errors": []}
    try:
        text = file.read().decode("utf-8-sig")
    except UnicodeDecodeError as e:
        raise ImportError_("file is not valid UTF-8 text") from e
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames or "ref_id" not in reader.fieldnames:
        raise ImportError_("missing header row with at least ref_id,result columns")
    entries = [
        entry
        for row in reader
        if any((value or "").strip() for value in row.values())
        and (entry := _entry(row, extras))
    ]
    return entries, extras


def parse_xlsx(file):
    from openpyxl import load_workbook

    extras = {"parse_errors": []}
    try:
        workbook = load_workbook(file, read_only=True, data_only=True)
    except Exception as e:
        raise ImportError_("file is not a valid xlsx workbook") from e
    sheet = workbook.worksheets[0]
    rows = sheet.iter_rows(values_only=True)
    try:
        header = [str(cell).strip().lower() if cell else "" for cell in next(rows)]
    except StopIteration:
        raise ImportError_("empty worksheet") from None
    if "ref_id" not in header:
        raise ImportError_("missing header row with at least ref_id,result columns")
    entries = []
    for values in rows:
        row = {
            header[i]: str(values[i]) if values[i] is not None else ""
            for i in range(min(len(header), len(values)))
        }
        if not any(value.strip() for value in row.values()):
            continue
        entry = _entry(row, extras)
        if entry:
            entries.append(entry)
    return entries, extras


def parse_ocsf(file):
    extras = {"parse_errors": [], "skipped_suppressed": 0, "skipped_other_class": 0}
    try:
        data = json.load(file)
    except (json.JSONDecodeError, UnicodeDecodeError) as e:
        raise ImportError_("file is not valid JSON") from e
    events = data if isinstance(data, list) else data.get("events")
    if not isinstance(events, list):
        raise ImportError_(
            "expected an array of OCSF events or an object with an 'events' array"
        )
    entries = []
    tool = ""
    for event in events:
        if not isinstance(event, dict):
            _record_parse_error(extras, event)
            continue
        if event.get("class_uid") != 2003:
            extras["skipped_other_class"] += 1
            continue
        if str(event.get("status", "")).lower() == "suppressed":
            extras["skipped_suppressed"] += 1
            continue
        compliance = event.get("compliance") or {}
        status = str(compliance.get("status", "")).lower()
        result = OCSF_STATUS_MAP.get(status) or OCSF_STATUS_ID_MAP.get(
            compliance.get("status_id")
        )
        if result is None:
            _record_parse_error(extras, {"status": compliance.get("status")})
            continue
        if not tool:
            product = (event.get("metadata") or {}).get("product") or {}
            tool = " ".join(
                filter(None, [product.get("name"), product.get("version")])
            )[:100]
        message = compliance.get("status_detail") or event.get("message") or ""
        for ref in compliance.get("requirements") or []:
            entries.append(
                {
                    "ref_id": str(ref).strip(),
                    "result": result,
                    "actual": "",
                    "expected": "",
                    "message": str(message)[:2000],
                }
            )
    if tool:
        extras["tool"] = tool
    return entries, extras


DELIMITERS = [",", ";", "\t", "|"]
AGGREGATIONS = ("worst_case", "best_case", "last_row", "strict")
WORST_ORDER = ["fail", "error", "not_checked", "pass", "not_applicable"]
BEST_ORDER = list(reversed(WORST_ORDER))
ANALYZE_VALUE_CAP = 20
IGNORE = "ignore"


def _decode(file):
    try:
        return file.read().decode("utf-8-sig")
    except UnicodeDecodeError as e:
        raise ImportError_("file is not valid UTF-8 text") from e


def _sniff_delimiter(header_line):
    counts = {d: header_line.count(d) for d in DELIMITERS}
    best = max(counts, key=counts.get)
    return best if counts[best] > 0 else ","


def analyze_csv(file, delimiter=None):
    text = _decode(file)
    first_line = text.split("\n", 1)[0]
    delimiter = delimiter or _sniff_delimiter(first_line)
    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
    headers = [h for h in (reader.fieldnames or []) if h]
    if not headers:
        raise ImportError_("missing header row")
    distinct = {h: set() for h in headers}
    sample_rows = []
    row_count = 0
    for row in reader:
        if not any((value or "").strip() for value in row.values()):
            continue
        row_count += 1
        if len(sample_rows) < 5:
            sample_rows.append({h: (row.get(h) or "")[:120] for h in headers})
        for h in headers:
            if len(distinct[h]) <= ANALYZE_VALUE_CAP:
                distinct[h].add((row.get(h) or "").strip())
    return {
        "delimiter": delimiter,
        "headers": headers,
        "row_count": row_count,
        "sample_rows": sample_rows,
        "columns": {
            h: (
                {"distinct": len(values), "values": sorted(values)}
                if len(values) <= ANALYZE_VALUE_CAP
                else {"distinct": f"{ANALYZE_VALUE_CAP}+", "values": None}
            )
            for h, values in distinct.items()
        },
    }


def _validate_mapping(mapping, headers):
    columns = mapping.get("columns") or {}
    for required in ("ref_id", "result"):
        if not columns.get(required):
            raise ImportError_(f"mapping.columns.{required} is required")
    for key, column in columns.items():
        if column and column not in headers:
            raise ImportError_(f"mapped column not in file: {column}")
    values = mapping.get("values") or {}
    if not isinstance(values, dict) or not values:
        raise ImportError_("mapping.values is required")
    for raw, target in values.items():
        if target not in RESULT_VALUES and target != IGNORE:
            raise ImportError_(f"invalid result binding for '{raw}': {target}")
    aggregation = mapping.get("aggregation", "worst_case")
    if aggregation not in AGGREGATIONS:
        raise ImportError_(f"unknown aggregation strategy: {aggregation}")
    return columns, values, aggregation


def _aggregate(rows, aggregation, key):
    if len(rows) == 1:
        return rows[0]
    if aggregation == "strict":
        raise ImportError_(f"multiple rows for ref_id '{key}' (strict aggregation)")
    if aggregation == "last_row":
        winner_row = rows[-1]
        winner = winner_row["result"]
    else:
        order = WORST_ORDER if aggregation == "worst_case" else BEST_ORDER
        present = {row["result"] for row in rows}
        winner = next(result for result in order if result in present)
        winner_row = next(row for row in rows if row["result"] == winner)
    winners = [row for row in rows if row["result"] == winner]
    message = next((row["message"] for row in winners if row["message"]), "")
    return {
        "ref_id": winner_row["ref_id"],
        "result": winner,
        "actual": f"{len(winners)}/{len(rows)} rows {winner}",
        "expected": winner_row["expected"],
        "message": message,
    }


def parse_mapped_csv(file, mapping):
    text = _decode(file)
    delimiter = mapping.get("delimiter") or _sniff_delimiter(text.split("\n", 1)[0])
    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
    headers = reader.fieldnames or []
    columns, values, aggregation = _validate_mapping(mapping, headers)

    extras = {"parse_errors": [], "skipped_ignored": 0, "skipped_unmapped": 0}
    asset_column = columns.get("asset")
    groups = {}
    for row in reader:
        if not any((value or "").strip() for value in row.values()):
            continue
        ref_id = (row.get(columns["ref_id"]) or "").strip()
        raw_result = (row.get(columns["result"]) or "").strip()
        if not ref_id:
            _record_parse_error(extras, {"row": str(row)[:200]})
            continue
        if raw_result not in values:
            extras["skipped_unmapped"] += 1
            continue
        result = values[raw_result]
        if result == IGNORE:
            extras["skipped_ignored"] += 1
            continue
        asset_value = (row.get(asset_column) or "").strip() if asset_column else ""

        def col(key):
            column = columns.get(key)
            return (row.get(column) or "").strip()[:2000] if column else ""

        groups.setdefault(asset_value, {}).setdefault(ref_id, []).append(
            {
                "ref_id": ref_id,
                "result": result,
                "actual": col("actual")[:255],
                "expected": col("expected")[:255],
                "message": col("message"),
            }
        )

    aggregated = {
        asset_value: [
            _aggregate(rows, aggregation, ref_id) for ref_id, rows in cells.items()
        ]
        for asset_value, cells in groups.items()
    }
    return aggregated, extras


PARSERS = {".csv": parse_csv, ".xlsx": parse_xlsx, ".json": parse_ocsf}


def parse_file(file):
    name = (file.name or "").lower()
    for suffix, parser in PARSERS.items():
        if name.endswith(suffix):
            return parser(file)
    raise ImportError_("unsupported file type (expected .csv, .xlsx or .json)")
